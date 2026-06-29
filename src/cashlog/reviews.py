"""App review crawling and Excel export helpers."""

from __future__ import annotations

import json
import re
import time
from copy import copy
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen


@dataclass
class ReviewRecord:
    source: str
    userName: str
    review: str
    score: int | float | None
    date: str | None
    title: str | None = None
    appVersion: str | None = None
    reviewId: str | None = None
    thumbsUpCount: int | None = None
    developerReply: str | None = None
    repliedAt: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def fetch_playstore_reviews(
    app_id: str,
    lang: str = "ko",
    country: str = "kr",
    target_count: int = 500,
    score: int | None = None,
) -> list[ReviewRecord]:
    """Fetch Google Play reviews newest-first."""

    try:
        from google_play_scraper import Sort, reviews
    except ImportError as exc:
        raise RuntimeError(
            "google-play-scraper is required. Install it with `pip install -e .[reviews]`."
        ) from exc

    total: list[dict[str, Any]] = []
    continuation_token = None
    page_size = min(200, max(1, target_count))

    while len(total) < target_count:
        batch, continuation_token = reviews(
            app_id,
            lang=lang,
            country=country,
            sort=Sort.NEWEST,
            count=min(page_size, target_count - len(total)),
            filter_score_with=score,
            continuation_token=continuation_token,
        )
        if not batch:
            break
        total.extend(batch)
        if continuation_token is None:
            break

    return [_normalize_playstore_review(row) for row in total[:target_count]]


def resolve_playstore_app_id(app_name: str, lang: str = "ko", country: str = "kr") -> str:
    """Resolve a Google Play package id from an app name."""

    try:
        from google_play_scraper import search
    except ImportError as exc:
        raise RuntimeError(
            "google-play-scraper is required. Install it with `pip install -e .[reviews]`."
        ) from exc

    results = search(app_name, lang=lang, country=country, n_hits=5)
    if results and results[0].get("appId"):
        return str(results[0]["appId"])

    web_app_id = _resolve_playstore_app_id_from_web(app_name, lang=lang, country=country)
    if web_app_id:
        return web_app_id

    for result in results:
        if result.get("appId"):
            return str(result["appId"])

    raise RuntimeError(f"No Play Store app found for {app_name!r} in country={country!r}.")


def fetch_appstore_reviews(
    app_name: str | None = None,
    country: str = "kr",
    target_count: int = 500,
    app_id: int | None = None,
    sleep: int | None = None,
) -> list[ReviewRecord]:
    """Fetch Apple App Store reviews."""

    if app_id is None and not app_name:
        raise ValueError("Pass app_name or app_id to fetch App Store reviews.")

    resolved_app_id = app_id or resolve_appstore_app_id(app_name or "", country)
    records: list[ReviewRecord] = []

    for page in range(1, 11):
        url = (
            f"https://itunes.apple.com/{country}/rss/customerreviews/page={page}/"
            f"id={resolved_app_id}/sortby=mostrecent/json"
        )
        feed = _read_json(url).get("feed", {})
        entries = feed.get("entry", [])
        if isinstance(entries, dict):
            entries = [entries]
        page_records = [_normalize_appstore_review(row) for row in entries if "im:rating" in row]
        if not page_records:
            break
        records.extend(page_records)
        if len(records) >= target_count:
            break
        if sleep:
            time.sleep(sleep)

    return records[:target_count]


def resolve_appstore_app_id(app_name: str, country: str = "kr") -> int:
    """Resolve an App Store numeric app id from an app name."""

    query = urlencode({"term": app_name, "country": country, "entity": "software", "limit": 1})
    payload = _read_json(f"https://itunes.apple.com/search?{query}")
    results = payload.get("results", [])
    if not results:
        raise RuntimeError(f"No App Store app found for {app_name!r} in country={country!r}.")
    return int(results[0]["trackId"])


def export_reviews_xlsx(
    reviews: list[ReviewRecord],
    output_path: str | Path,
    sheet_name: str = "reviews",
) -> Path:
    """Write normalized reviews to an Excel workbook."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install it with `pip install -e .[reviews]`.") from exc

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "reviews"

    headers = [
        "source",
        "userName",
        "review",
        "score",
        "date",
        "title",
        "appVersion",
        "reviewId",
        "thumbsUpCount",
        "developerReply",
        "repliedAt",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for record in reviews:
        row = record.to_dict()
        ws.append([row.get(header) for header in headers])

    widths = {
        "A": 12,
        "B": 22,
        "C": 70,
        "D": 10,
        "E": 14,
        "F": 36,
        "G": 16,
        "H": 38,
        "I": 16,
        "J": 50,
        "K": 20,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, ws.max_row)}"

    wb.save(output_path)
    return output_path


def _normalize_playstore_review(row: dict[str, Any]) -> ReviewRecord:
    return ReviewRecord(
        source="PlayStore",
        userName=str(row.get("userName") or ""),
        review=str(row.get("content") or ""),
        score=row.get("score"),
        date=_format_date(row.get("at")),
        appVersion=_optional_str(row.get("appVersion") or row.get("reviewCreatedVersion")),
        reviewId=_optional_str(row.get("reviewId")),
        thumbsUpCount=row.get("thumbsUpCount"),
        developerReply=_optional_str(row.get("replyContent")),
        repliedAt=_format_date(row.get("repliedAt")),
    )


def _normalize_appstore_review(row: dict[str, Any]) -> ReviewRecord:
    title = _label(row.get("title"))
    body = _label(row.get("content")) or ""
    review = f"{title} {body}".strip() if title else body
    return ReviewRecord(
        source="AppStore",
        userName=_label(row.get("author", {}).get("name")) or "",
        review=review,
        score=_parse_int(_label(row.get("im:rating"))),
        date=_format_date(_label(row.get("updated"))),
        title=title,
        appVersion=_label(row.get("im:version")),
        reviewId=_label(row.get("id")),
    )


def _format_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    return str(value)


def _label(value: Any) -> str | None:
    if isinstance(value, dict) and "label" in value:
        return _optional_str(value.get("label"))
    return _optional_str(value)


def _parse_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _read_json(url: str) -> dict[str, Any]:
    request = Request(url, headers={"User-Agent": "cashlog-auto/0.1"})
    with urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def _resolve_playstore_app_id_from_web(
    app_name: str, lang: str = "ko", country: str = "kr"
) -> str | None:
    query = urlencode({"q": app_name, "c": "apps", "hl": lang, "gl": country})
    request = Request(
        f"https://play.google.com/store/search?{query}",
        headers={"User-Agent": "Mozilla/5.0"},
    )
    with urlopen(request, timeout=30) as response:
        html = response.read().decode("utf-8", errors="ignore")

    seen: set[str] = set()
    for app_id in re.findall(r"/store/apps/details\?id=([A-Za-z0-9._]+)", html):
        if app_id not in seen:
            return app_id
        seen.add(app_id)
    return None
